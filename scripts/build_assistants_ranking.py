"""Gera docs/product/assistants-ranking.xlsx — ranking esforço×impacto do
pipeline de assistentes/funcionalidades do PROGPT.

Fonte dos dados: a avaliação em
C:/Users/.../.claude/plans/avalie-este-pipeline-*.md (lente: lançar rápido /
validar receita). Score = Impacto / Esforço. Entregues no topo; construíveis
ranqueados por score desc (desempate por impacto desc).

Rodar:  python scripts/build_assistants_ranking.py
"""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parent.parent / "docs" / "product" / "assistants-ranking.xlsx"

# --- dados -----------------------------------------------------------------
# (ideia, categoria, passo, status, impacto, esforco, veredito, motivo)
# status: "existe" | "redundante" | "novo"
SHIPPED = [
    ("Profile (Perfil da Categoria)", "Entregue", "1", "existe", None, None, "Entregue", "Em produção"),
    ("ABC / Pareto", "Entregue", "2", "existe", None, None, "Entregue", "Em produção"),
    ("Porter (5 Forças)", "Entregue", "3", "existe", None, None, "Entregue", "Em produção"),
    ("Supplier Search (CNAE/região)", "Entregue", "3", "existe", None, None, "Entregue", "Em produção"),
    ("Kraljic (Matriz)", "Entregue", "4", "existe", None, None, "Entregue", "Em produção"),
    ("RFP / RFQ", "Entregue", "5", "existe", None, None, "Entregue", "Em produção"),
    ("Negotiation Simulator", "Entregue", "6", "existe", None, None, "Entregue", "Em produção"),
    ("Financial Score", "Entregue", "7", "existe", None, None, "Entregue", "Em produção"),
]

BUILDABLE = [
    ("Supplier Scorecard", "Planejado T1", "8", "novo", 4, 3, "Construir 1º",
     "Melhor bang-for-buck; reusa padrão Kraljic; SRM defensável"),
    ("Contract Clause Generator", "Planejado T1", "7", "novo", 5, 4, "Construir cedo",
     "Maior dor declarada (semanal); de-riscar jurídico (v1 + disclaimer)"),
    ("Supplier Segmentation Matrix", "Planejado T3", "4", "novo", 3, 3, "Adiar",
     "Distinto de Kraljic, mas risco de confusão; baixa urgência de venda"),
    ("Artifact Versioning (v1/v2/v3)", "Cross-cutting", "—", "novo", 3, 3, "Multiplicador",
     "Aumenta valor dos 8 sem tile novo; pós-validação"),
    ("Category Profile Generator", "Planejado T3", "1", "redundante", 2, 2, "Não fazer",
     "Redundante com Profile; se preciso, vira modo do Profile"),
    ("RFI Generator", "Planejado T3", "5", "redundante", 2, 2, "Não fazer",
     "Subconjunto do RFP; virar toggle 'RFI' no RFP, não 9º tile"),
    ("PDCA Cycle Helper", "Backlog", "—", "novo", 2, 2, "Baixa prioridade",
     "Nicho; baixa urgência"),
    ("Contract Risk Review", "Planejado T3", "7", "novo", 4, 5, "Adiar",
     "Alto valor mas heavy (file ingestion) + risco jurídico alto"),
    ("TCO Calculator", "Planejado T2", "3", "novo", 3, 4, "Adiar",
     "Audiência madura/estreita; não é driver de fechamento"),
    ("Savings Tracker", "Backlog", "—", "novo", 3, 4, "Pós-validação",
     "CFO-facing; bom pra retenção/expansão, não pra 1ª venda"),
    ("ESG / Sustainability Assessment", "Backlog", "—", "novo", 3, 4, "Sob demanda",
     "Pressão regulatória crescente; construir quando cliente pedir"),
    ("Maverick Spend Detector", "Backlog", "—", "novo", 3, 4, "Sob demanda",
     "Valor de compliance; precisa parsing de spend"),
    ("Template Library por org", "Cross-cutting", "—", "novo", 3, 4, "Fora da lente",
     "Feature B2B (multi-tenancy); fora do foco B2C launch"),
    ("Negotiation Strategy Brief", "Planejado T2", "6", "redundante", 2, 3, "Não fazer",
     "Redundante com o Strategy Builder do Negotiation Simulator"),
    ("Make-vs-Buy Analysis", "Backlog", "—", "novo", 2, 3, "Baixa prioridade",
     "Nicho; baixa urgência"),
    ("Notifications", "Cross-cutting", "—", "novo", 2, 3, "Fora da lente",
     "Depende de sharing (B2B); fora do foco"),
    ("Risk Score Aggregator", "Backlog", "—", "novo", 3, 5, "Não fazer",
     "Depende de dado externo (D&B-like) caro; demanda incerta"),
    ("Should-Cost Modeling", "Backlog", "—", "novo", 3, 5, "Não fazer",
     "Nicho + pesado (decomposição de custo data-heavy)"),
    ("Workspace Sharing", "Cross-cutting", "—", "novo", 3, 5, "Fora da lente",
     "Feature B2B (multi-tenancy); fora do foco B2C launch"),
    ("Supplier Onboarding Wizard", "Backlog", "—", "novo", 2, 4, "Não fazer",
     "Workflow/KYC, não análise; território de ERP/suite"),
]


def score(impacto, esforco):
    return round(impacto / esforco, 2)


# ordena construíveis: score desc, depois impacto desc
buildable_sorted = sorted(
    BUILDABLE, key=lambda r: (score(r[4], r[5]), r[4]), reverse=True
)

# --- estilos ---------------------------------------------------------------
GREEN = PatternFill("solid", fgColor="C6EFCE")
RED = PatternFill("solid", fgColor="FFC7CE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
BLUE = PatternFill("solid", fgColor="BDD7EE")
GRAY = PatternFill("solid", fgColor="D9D9D9")
HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
BORDER = Border(*(Side(style="thin", color="D0D0D0"),) * 4)


def veredito_fill(veredito, status):
    if status == "existe":
        return GREEN
    if veredito.startswith("Construir") or veredito == "Multiplicador":
        return BLUE
    if veredito == "Adiar":
        return YELLOW
    if veredito == "Não fazer":
        return RED
    return GRAY  # pós-validação / sob demanda / baixa prioridade / fora da lente


def status_label(status):
    return {"existe": "✅ Já existe", "redundante": "🔁 Redundante", "novo": "🆕 Novo"}[status]


# --- planilha --------------------------------------------------------------
wb = Workbook()
ws = wb.active
ws.title = "Ranking"

headers = ["#", "Ideia", "Categoria", "Passo SS", "Status",
           "Impacto", "Esforço", "Score", "Veredito", "Motivo"]
ws.append(headers)
for c in range(1, len(headers) + 1):
    cell = ws.cell(row=1, column=c)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

row_idx = 2


def write_row(rank, idea, cat, passo, status, impacto, esforco, veredito, motivo):
    global row_idx
    sc = "" if impacto is None else score(impacto, esforco)
    esf = "Feito" if esforco is None else esforco
    imp = "" if impacto is None else impacto
    values = [rank, idea, cat, passo, status_label(status), imp, esf, sc, veredito, motivo]
    for c, v in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.border = BORDER
        cell.alignment = Alignment(
            horizontal="center" if c in (1, 4, 5, 6, 7, 8) else "left",
            vertical="center",
            wrap_text=(c in (2, 9, 10)),
        )
    # cor no Status e no Veredito
    fill = veredito_fill(veredito, status)
    ws.cell(row=row_idx, column=5).fill = GREEN if status == "existe" else (RED if status == "redundante" else PatternFill())
    ws.cell(row=row_idx, column=9).fill = fill
    row_idx += 1


# entregues primeiro
for r in SHIPPED:
    write_row("✓", *r)

# construíveis ranqueados
for i, r in enumerate(buildable_sorted, start=1):
    write_row(i, *r)

# larguras
widths = [5, 32, 14, 9, 14, 9, 9, 8, 16, 52]
for i, w in enumerate(widths, start=1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.freeze_panes = "A2"
ws.row_dimensions[1].height = 30

# nota de rodapé
note_row = row_idx + 1
ws.cell(row=note_row, column=1,
        value=("Lente: lançar rápido / validar receita. Score = Impacto ÷ Esforço (1-5). "
               "Recomendação: não construir assistente novo antes de validar os 8 atuais "
               "com pagantes; instrumentar uso e deixar os dados escolherem o próximo."))
ws.cell(row=note_row, column=1).font = Font(italic=True, size=9, color="666666")
ws.merge_cells(start_row=note_row, start_column=1, end_row=note_row, end_column=10)

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"OK: {OUT}  ({len(SHIPPED)} entregues + {len(buildable_sorted)} construíveis)")
